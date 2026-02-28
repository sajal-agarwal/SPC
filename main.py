import openpyxl
import pandas
import pandas as pd
import threading
import math
import shutil
import xlwings as xlwings
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Constants
COLUMN_WIDTH = 16
INDEX_COLUMN_WIDTH = 8
NAME_COLUMN_WIDTH = 24
HEADER_ROW_INDEX = 1

pd.options.mode.chained_assignment = None
# pd.set_option('display.max_columns', None)
# pd.set_option('display.max_rows', None)

df = pd.DataFrame()
untouched_df = pd.DataFrame()
summary_df = pd.DataFrame()
columns = []
last_err_str: str = ''
progress = 0
exit_flag = False
df_updated = False
remove_if_str: str = ''
include_if_str: str = ''
rules = {}
sheet_columns = []
highlight_columns = []

progress_lock = threading.Lock()
error_lock = threading.Lock()
exit_flag_lock = threading.Lock()
df_update_lock = threading.Lock()


def set_progress(cur_progress):
    with progress_lock:
        global progress
        progress = cur_progress


def get_progress():
    with progress_lock:
        return progress


def set_last_error(cur_err):
    with error_lock:
        global last_err_str
        last_err_str = cur_err


def get_last_error() -> str:
    with error_lock:
        return last_err_str


def set_exit_flag(flag):
    with exit_flag_lock:
        global exit_flag
        exit_flag = flag


def get_exit_flag():
    with exit_flag_lock:
        return exit_flag


def set_df_updated(updated):
    with df_update_lock:
        global df_updated
        df_updated = updated


def get_df_updated():
    with df_update_lock:
        return df_updated


def get_columns():
    return columns


def get_avg_columns():
    return avg_cols


avg_cols = []


def set_average_cols(cols):
    global avg_cols
    avg_cols = [columns[col] for col in cols]


def set_average_cols_str(cols):
    global avg_cols
    avg_cols = cols


del_cols = []


def set_deleted_cols(cols):
    global del_cols
    del_cols = [columns[col] for col in cols]


def set_deleted_cols_str(cols):
    global del_cols
    del_cols = cols


def set_sheet_cols_str(cols):
    global sheet_columns
    sheet_columns = cols


def get_deleted_cols():
    global del_cols
    return del_cols


def get_remove_if_str():
    global remove_if_str
    return remove_if_str


def set_remove_if_str(rem_str):
    global remove_if_str
    remove_if_str = rem_str


def get_include_if_str():
    global include_if_str
    return include_if_str


def set_include_if_str(add_str):
    global include_if_str
    include_if_str = add_str


def set_rules(r):
    global rules
    rules = r


def get_rules():
    global rules
    return rules


def get_df():
    global df
    return df


def set_sheet_columns(cols):
    global columns
    global sheet_columns
    sheet_columns = [columns[col] for col in cols]


def get_sheet_columns():
    global sheet_columns
    return sheet_columns


def set_sheet_highlight_columns(cols):
    global columns
    global highlight_columns
    highlight_columns = [columns[col] for col in cols]


def get_sheet_highlight_columns():
    global highlight_columns
    return highlight_columns


def clear():
    global df
    global summary_df
    global columns
    global last_err_str
    global progress
    global exit_flag
    global df_updated
    global avg_cols
    global del_cols
    # global remove_if_str
    # global include_if_str

    df = pd.DataFrame()
    summary_df = pd.DataFrame()
    set_last_error('')
    set_progress(0)
    set_exit_flag(False)
    set_df_updated(False)
    columns = []
    avg_cols = []
    del_cols = []
    # remove_if_str = ''
    # include_if_str = ''


def is_int(s):
    try:
        int(s)
        return True
    except ValueError:
        return False


def is_float(s):
    try:
        float(s)
        return True
    except ValueError:
        return False


def is_str_numeric(s):
    return is_int(s) or is_float(s)


def update_average(cl_df, col_name):
    cl_df.at['Average', col_name] = round(cl_df[col_name].mean(), 2)


def is_val_exits_in_column(col, val):
    for item in df[col]:
        if val == str(item):
            return True
    return False


def is_column_numeric(col):
    global df
    is_numeric = True
    count_nan = 0
    for item in df[col]:
        if not isinstance(item, (int, float)):
            is_numeric = False
            break
        elif math.isnan(item):
            count_nan += 1

    faulty_row = -1
    row_index = 0
    if not is_numeric:
        for item in df[col]:
            if (not isinstance(item, (int, float))) and (not isinstance(item, str) or (not is_str_numeric(item))):
                faulty_row = row_index + 2  # +1 for header, +1 to convert row index to row number
                break

            row_index += 1

    return is_numeric and (count_nan < df[col].size), faulty_row


def select_all_numeric_cols(cur_tab):
    apply_rules()
    if get_df_updated():
        global avg_cols
        global highlight_columns
        if cur_tab == 2:
            avg_cols.clear()
        elif cur_tab == 7:
            highlight_columns.clear()
        for col in columns:
            if is_column_numeric(col)[0]:
                if cur_tab == 2:
                    avg_cols.append(col)
                elif cur_tab == 7:
                    highlight_columns.append(col)


tmp_col_rules = {}


def get_val_from_rule(str1):
    if isinstance(str1, float) and math.isnan(str1):
        str1 = str(str1)

    val = tmp_col_rules.get(str1, '')
    if len(val) > 0:
        if is_int(val):
            return int(val)
        elif is_float(val):
            return float(val)
        else:
            return val
    else:
        return str1


def apply_rules_on_column(df1, col_name):
    if len(rules) == 0:
        return

    if col_name in rules:
        global tmp_col_rules
        tmp_col_rules = rules[col_name]
        df1[col_name] = df[col_name].apply(get_val_from_rule)


def apply_rules():
    reset_df()
    if get_df_updated() and (len(rules) > 0):
        global df
        df.reset_index(inplace=True, drop=True)
        for col in columns:
            apply_rules_on_column(df, col)


def update_df(in_paths):
    success = True
    set_last_error('')
    set_df_updated(False)
    global df
    global untouched_df
    try:
        untouched_df = pd.DataFrame()
        for in_path in in_paths:
            untouched_df = pandas.concat([untouched_df, pd.read_excel(in_path)])

        global columns
        columns = untouched_df.columns.to_list()
        set_df_updated(True)
    except Exception as e:
        set_last_error(e)
        success = False

    df = untouched_df.copy()

    apply_rules()

    return success


def reset_df():
    global df
    global untouched_df
    df = untouched_df.copy()


def delete_columns():
    global del_cols
    for col in del_cols:
        if col in df:
            df.drop(col, inplace=True, axis=1)


def create_output_file_from_template(out_file):
    try:
        shutil.copyfile('./data/OutputTemplate.xlsx', out_file)
    except PermissionError as _:
        book = xlwings.Book(out_file)
        book.close()

        shutil.copyfile('./data/OutputTemplate.xlsx', out_file)


def get_count_df():
    global df
    global avg_cols
    if get_df_updated():
        count_df = pd.DataFrame(columns=avg_cols, index=range(10, 0, -1))
        count_df.index.name = 'Points'
        for col in avg_cols:
            vc = df[col].value_counts()
            count_df[col] = [(vc[i] if i in vc else 0) for i in range(10, 0, -1)]
        return count_df
    return pd.DataFrame()


def do_remove_if():
    global df
    global remove_if_str

    if len(remove_if_str) == 0:
        return

    conditions = remove_if_str.split('\n')
    try:
        for cond in conditions:
            if cond == '':
                continue

            if '==' in cond:
                cond_list = cond.split('==')
                if len(cond_list) == 0:
                    raise Exception("")
                df = df[[(str(i) != cond_list[1]) for i in df[cond_list[0]]]]
            else:
                raise Exception('')
            # elif '!=' in cond:
            #     cond_list = cond.split('!=')
            #     if len(cond_list) == 0:
            #         continue
            #     new_df.append(df[[(str(i) == cond_list[1]) for i in df[cond_list[0]]]])
            # elif '<' in cond:
            #     cond_list = cond.split('<')
            #     if len(cond_list) == 0:
            #         continue
            #     df = df[df[cond_list[0]] < cond_list[1]]
            # elif '<=' in cond:
            #     cond_list = cond.split('<=')
            #     if len(cond_list) == 0:
            #         continue
            #     df = df[df[cond_list[0]] <= cond_list[1]]
            # elif '>' in cond:
            #     cond_list = cond.split('>')
            #     if len(cond_list) == 0:
            #         continue
            #     df = df[df[cond_list[0]] > cond_list[1]]
            # elif '>=' in cond:
            #     cond_list = cond.split('>=')
            #     if len(cond_list) == 0:
            #         continue
            #     df = df[df[cond_list[0]] >= cond_list[1]]
    except Exception as e:
        raise Exception("Invalid condition - " + ("only == supported with no extra spaces"
                                                  if (len(str(e)) == 0) else str(e)))


def do_include_if():
    global df
    global include_if_str

    if len(include_if_str) == 0:
        return

    conditions = include_if_str.split('\n')
    new_df = pd.DataFrame()
    try:
        for cond in conditions:
            if cond == '':
                continue

            if '==' in cond:
                cond_list = cond.split('==')
                if len(cond_list) == 0:
                    raise Exception("")

                new_df = pandas.concat([new_df, df[[(str(i) == cond_list[1]) for i in df[cond_list[0]]]]])
            else:
                raise Exception('')
        df = new_df
    except Exception as e:
        raise Exception("Invalid condition - " + ("only == supported with no extra spaces"
                                                  if (len(str(e)) == 0) else str(e)))


def get_cell_width(cell):
    if cell.value is None or len(cell.value) == 0:
        return INDEX_COLUMN_WIDTH
    elif cell.value.startswith('Name of the child'):
        return NAME_COLUMN_WIDTH
    else:
        return COLUMN_WIDTH


def format_tabel_header_in_all_sheets(workbook):
    for sheet_name in workbook.sheetnames:
        if sheet_name == 'Graphical Representation':
            continue

        sheet = workbook[sheet_name]
        for cell in sheet[HEADER_ROW_INDEX]:
            column_letter = get_column_letter(cell.column)
            sheet.column_dimensions[column_letter].width = get_cell_width(cell)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.fill = PatternFill(patternType='solid', fgColor='C6E0B4')

        sheet.auto_filter.ref = sheet.dimensions


def format_average_row_in_all_sheets(workbook):
    for sheet_name in workbook.sheetnames:
        if sheet_name == 'Graphical Representation' or sheet_name == 'MasterSheet':
            continue

        sheet = workbook[sheet_name]
        row_index = 2
        average_row_found = False
        while sheet[row_index][0] is not None and sheet[row_index][0].value is not None:
            if sheet[row_index][0].value == 'Average' or sheet[row_index][0].value == 'Weighted Average':
                average_row_found = True
                break
            row_index += 1

        if average_row_found:
            for cell in sheet[row_index]:
                cell.border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))


def set_row_color(worksheet, row_number, color):
    for cell1 in worksheet[row_number]:
        cell1.fill = PatternFill(patternType='solid', fgColor=color)


def highlight_rows(workbook):
    for sheet_name in workbook.sheetnames:
        if sheet_name == 'Graphical Representation' or sheet_name == 'Counts':
            continue

        ws = workbook[sheet_name]
        row_index = 2
        for row in ws.iter_rows(min_row=2):
            set_background_color = False
            col_index = 0
            for cell in row:
                if ws[1][col_index].value in highlight_columns:
                    if isinstance(cell.value, (int, float)) and cell.value <= 7:
                        set_background_color = True
                    elif isinstance(cell.value, (str,)) and is_str_numeric(cell.value) and int(cell.value) <= 7:
                        set_background_color = True

                col_index += 1

            if set_background_color:
                set_row_color(ws, row_index, "FFFF00")

            row_index += 1


def do_work(in_paths, out_path):
    global df
    global summary_df
    success = True
    try:
        set_last_error('')

        if not get_df_updated():
            update_df(in_paths)
        else:
            apply_rules()

        delete_columns()

        do_remove_if()

        do_include_if()

        create_output_file_from_template(out_path)

        with pd.ExcelWriter(out_path, mode="a") as writer:
            df.to_excel(writer, sheet_name='MasterSheet', index=False)

        classes = []
        for col in sheet_columns:
            if len(classes) == 0:
                classes = df[col].astype(str).values.tolist()
                continue

            classes = ["{}-{}".format(a, b) for a, b in zip(classes, df[col])]

        classes = list(set(classes))
        classes.sort()

        is_summary_needed = (len(avg_cols) > 0)

        if is_summary_needed:
            summary_df = pd.DataFrame(columns=avg_cols, index=[str1.replace("-", "") for str1 in classes])

        max_count = (len(classes) + 2)
        # New dataframe to hold class name and feedback count
        df_class_feedback_count = pd.DataFrame(columns=['Sheet Name', 'Feedback Count'])
        counter = 0
        for c in classes:
            if get_exit_flag():
                break

            cl_sec = c.split('-')
            choices = pd.Series([True]*len(df), index=df.index)
            for col, val in zip(sheet_columns, cl_sec):
                choices &= (df[col] == val)

            cl_df = df[choices]
            cl_df.reset_index(inplace=True, drop=True)
            cl_df.index += 1
            for col in avg_cols:
                update_average(cl_df, col)

            row_name = ''
            for val in cl_sec:
                row_name += val

            if is_summary_needed:
                av_ser = cl_df.iloc[-1]
                av_ser = av_ser[~av_ser.isnull()]
                summary_df.loc[row_name] = av_ser

            with pd.ExcelWriter(out_path, mode="a") as writer:
                cl_df.to_excel(writer, sheet_name=row_name)
                df_class_feedback_count.loc[len(df_class_feedback_count)] = [row_name, len(cl_df) - 1]

            counter += 1
            set_progress(counter*100/max_count)

        if not get_exit_flag():
            count_df = get_count_df()
            if is_summary_needed:
                if len(classes) > 0:
                    avg_of_avg = summary_df.mean().astype(float).round(decimals=2)
                    count_df.loc['Weighted Average'] = avg_of_avg
                    summary_df.loc['Average'] = avg_of_avg

                    with pd.ExcelWriter(out_path, mode="a") as writer:
                        summary_df.to_excel(writer, sheet_name='Summary')

                else:
                    for col in avg_cols:
                        update_average(df, col)

                    count_df.loc['Weighted Average'] = df.loc['Average']
                    df.drop(df.index[-1], inplace=True)

            counter += 1
            set_progress(counter*100/max_count)

            with pd.ExcelWriter(out_path, mode="a") as writer:
                df_class_feedback_count.to_excel(writer, sheet_name='Feedback Count', index=False)

            with pd.ExcelWriter(out_path, mode="a", if_sheet_exists='overlay') as writer:
                count_df.to_excel(writer, sheet_name='Counts')

            workbook = openpyxl.load_workbook(filename=out_path)
            format_tabel_header_in_all_sheets(workbook)
            format_average_row_in_all_sheets(workbook)
            highlight_rows(workbook)
            workbook.save(out_path)
    except Exception as e:
        set_last_error(e)
        success = False

    set_progress(100)
    return success
